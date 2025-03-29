# Importing the required libraries for the project
import sys

# Adding the code directory to the path so that the code can be imported
sys.path.append('./code')
import email
import email.utils
import imaplib

import openpyxl as xl  # Importing the openpyxl library to work with Excel files
import yaml
from openpyxl import \
    Workbook  # Importing the Workbook class from the openpyxl library
from openpyxl.utils import \
    get_column_letter  # Importing the get_column_letter function from the openpyxl library to get the column letter from the column number


def write_data_to_excel(msgs):
    wb = Workbook() # Create a new Workbook object
    ws = wb.active # Get the active worksheet from the Workbook object
    ws.title = "Gmail Data" # Set the title of the worksheet to "Gmail Data"
    
    # #Now we have all messages, but with a lot of details
    # #Let us extract the right text  and write to the Excel file
    # # #In a multipart e-mail, email.message.Message.get_payload() returns a
    # # list with one item for each part. The easiest way is to walk the message
    # # and get the payload on each part:
    # # https://stackoverflow.com/questions/1463074/how-can-i-get-an-email-messages-text-content-using-python
    # # NOTE that a Message object consists of headers and payloads.

    # Write the header row to the Excel file
    ws.append(["Sender", "Date Received", "Subject", "Body"])
    
    #Set bold font for the header row
    for cell in ws[1]:
        cell.font = xl.styles.Font(bold=True)
        
    #Set the width of the header row
    for cell in ws[1]:
        cell.alignment = xl.styles.Alignment(horizontal='center')
        cell.fill = xl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        cell.border = xl.styles.Border(left=xl.styles.Side(style='thin'),
                                       right=xl.styles.Side(style='thin'),
                                       top=xl.styles.Side(style='thin'),
                                       bottom=xl.styles.Side(style='thin'))
        cell.font = xl.styles.Font(bold=True, color='000000')

    for msg in msgs[::-1]:
        for response_part in msg:
            if type(response_part) is tuple:
                my_msg=email.message_from_bytes((response_part[1]))

                # Get the sender of the email
                sender = my_msg['from']

                # Get the date of the email
                date = email.utils.parsedate_to_datetime(my_msg['date'])
                date = date.strftime("%Y-%m-%d %H:%M:%S")

                # Get the subject of the email
                subject = my_msg['subject']

                # Get the body of the email
                for part in my_msg.walk():  
                    #print(part.get_content_type())
                    if part.get_content_type() == 'text/plain':
                        # print (part.get_payload())
                        body = part.get_payload()
                        
                # Write the data to the Excel file
                ws.append([sender,  date, subject, body])

    # Set the column width to fit the content 
    for column in ws.columns:
        max_length = 0

        # # Get the maximum length of the content in the column
        # column = [cell for cell in column]
        # for cell in column:
        #     try:
        #         if len(str(cell.value)) > max_length:
        #             max_length = len(str(cell.value))
        #     except:
        #         pass

        # Set the column width to the maximum length of the content in the column
        try:
            max_length = max(len(str(cell.value)) for cell in column)
        except:
            pass

        # Adjust the width of the column to fit the content
        adjusted_width = (max_length + 2)
        # Set the width of the column to the adjusted width
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    
    # Save the Excel file
    wb.save("gmail_data.xlsx")